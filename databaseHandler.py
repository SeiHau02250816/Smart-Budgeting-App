import openpyxl
import zipfile
from transaction import Transaction
from datetime import datetime

class ExcelDatabase:
    def __init__(self, file_path):
        """
        Initialize the Excel database.

        :param file_path: Path to the Excel file.
        """
        self.file_path = file_path
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.transactions = self.get_data_from_sheet('Spent')  # Load existing transactions
        except (FileNotFoundError, openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile):
            self.workbook = openpyxl.Workbook()
            self.workbook.save(file_path)
            self.transactions = []  # Initialize empty transactions list if file is new or invalid
        
    def get_sheet_names(self):
        """
        Get a list of sheet names in the Excel file.

        :return: List of sheet names.
        """
        return self.workbook.sheetnames

    def create_sheet(self, sheet_name):
        """
        Create a new sheet in the Excel file.

        :param sheet_name: Name of the new sheet.
        """
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.file_path)

    def get_data_from_sheet(self, sheet_name):
        """
        Retrieve data from a specific sheet and return as a list of Transaction objects.

        :param sheet_name: Name of the sheet to retrieve data from.
        :return: List of Transaction objects.
        """
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            transactions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
                transaction = Transaction(
                    date=row[1].strftime("%Y-%m-%d") if isinstance(row[1], datetime) else row[1],
                    amount=row[2],
                    business_name=row[3],
                    category=row[4]
                )
                transactions.append(transaction)
            return transactions
        return None

    def add_data_to_sheet(self, sheet_name, transactions):
        """
        Add transaction data to a specific sheet.

        :param sheet_name: Name of the sheet to add data to.
        :param transactions: List of Transaction objects to add.
        """
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for transaction in transactions:
                sheet.append([transaction.transaction_id, transaction.date, transaction.amount, transaction.business_name, transaction.category.value])
                self.transactions.append(transaction)  # Record the transaction
            self.workbook.save(self.file_path)

    def create_spent_table(self):
        """
        Create a 'Spent' table with specified attributes in the Excel file.
        """
        if 'Spent' not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(title='Spent')
            headers = ['Txn ID', 'Date', 'Amount (RM)', 'Business Name', 'Transaction Category']
            sheet.append(headers)
            self.workbook.save(self.file_path)

    def print_spent_data(self):
        """
        Print data from the 'Spent' table row by row.
        """
        data = self.get_data_from_sheet('Spent')
        if data:
            for row in data:
                print(row)

    def get_all_transactions(self):
        """
        Return all transactions stored in the database.

        :return: List of Transaction objects.
        """
        return self.transactions

    def get_total_spending(self):
        """
        Calculate the sum of all transaction amounts.

        :return: Total amount of all transactions.
        """
        total = 0
        for transaction in self.transactions:
            total += transaction.amount
        return total

if __name__ == "__main__":
    db = ExcelDatabase('database.xlsx')
    db.create_spent_table()
    db.add_data_to_sheet('Spent', transactions)
    db.print_spent_data()